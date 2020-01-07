
import sys, os
import math
from openpyxl import load_workbook

from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QMessageBox, QVBoxLayout, QHBoxLayout, QDialog, QMessageBox
from PyQt5.QtGui import QPixmap, QIcon, QImage
from PyQt5 import QtWidgets, QtGui




def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


#--------------------------------------
caminho = resource_path('tabela_gerdau.xlsx')
arquivo_excel = load_workbook(caminho)

planilha = arquivo_excel['Plan1']
mlinhas = planilha.max_row
mcolunas = planilha.max_column
lista = []

for i in range(5,mlinhas+1):
	aux_bitola = planilha.cell(i,1).value
	lista.append(aux_bitola)

# ------
planilha_acos = arquivo_excel['Plan2']
mlinhas_acos = planilha_acos.max_row
mcolunas_acos = planilha_acos.max_column
lista_acos = []

for i in range(4,mlinhas_acos+1):
	aux_bitola = planilha_acos.cell(i,1).value
	lista_acos.append(aux_bitola)

lista_apoios = ['Engaste A','Engaste B','Engaste C','Engaste D','Engaste E','Engaste F']
#--------------------------------------


class Aplicacao(QMainWindow):
	def __init__(self):
		super().__init__()
		self.ui = None
		self.load_ui()
		self.load_signals()
		self.setWindowTitle('BRUNEL - verificação de barras metálicas laminadas perfil I e H')
		self.setWindowIcon(QtGui.QIcon(resource_path(resource_path('imagens/icone_brunel.png'))))

		self.setFixedSize(590, 495)

	def load_ui(self):
		self.ui = loadUi(resource_path('main_ui.ui'),self)

		self.carregar_init_dados()

		self.comboBox.addItems(lista)
		self.comboBox_2.addItems(lista_acos)
		self.comboBox_3.addItems(lista_apoios)

		self.label_93.setStyleSheet("Background-Color: #ddebff;")
		self.label_94.setToolTip('LAEC - Liga Acadêmica de Engenharia Civil \n/ Centro Acadêmico Alberto Silva')
		self.label_95.setToolTip('Isambard Kingdom Brunel (1806-1859), \nfoi engenheiro, arquiteto e inventor. \nConsiderado um dos que maiores \nnomes do séc. XIX.')  

		#self.pushButton_8.setStyleSheet("Background: url('logo_info_mini.png')")

		self.show()

	def load_signals(self):
		self.comboBox.currentIndexChanged.connect(self.dados_perfil)
		self.comboBox_2.currentIndexChanged.connect(self.dados_aco)

		self.pushButton.clicked.connect(self.abaEsquemaApoios)
		self.pushButton_9.clicked.connect(self.abaEsquemaCb)

		self.comboBox_3.currentIndexChanged.connect(self.altera_K)
		self.pushButton_2.clicked.connect(self.calcular_normal)
		self.pushButton_3.clicked.connect(self.limpa_normal)
		self.pushButton_4.clicked.connect(self.salvar_normal)

		self.pushButton_5.clicked.connect(self.calcular_flt)
		self.pushButton_5.clicked.connect(self.calcular_flm)
		self.pushButton_5.clicked.connect(self.calcular_fla)
		self.pushButton_6.clicked.connect(self.limpa_momento)
		self.pushButton_7.clicked.connect(self.salvar_momento)

		self.checkBox.clicked.connect(self.alterna_kv)
		self.pushButton_15.clicked.connect(self.calcular_cortante)
		self.pushButton_17.clicked.connect(self.limpar_cortante)
		self.pushButton_16.clicked.connect(self.salvar_cortante)

		#self.pushButton_8.clicked.connect(self.abaInformacao)
		self.pushButton_8.clicked.connect(self.info_central)


	def info_central(self):
		msg = QMessageBox()
		msg.setWindowIcon(QtGui.QIcon(resource_path("imagens/icone_brunel.png")))

		pixmap = QPixmap(resource_path("imagens/logo_a3_mini_l.png"))
		msg.setIconPixmap(pixmap)
		#msg.setIcon(QMessageBox.Information)
		msg.setText('BRUNEL <i>    -v. beta</i>at. 07.2019 <hr> Desenvolvido por Anderson Alves de Aguiar<br> (Acadêmico de Engenharia Civil)<br>- andersonalvesmath@hotmail.com')
		msg.setWindowTitle("Informação")
		#msg.setDetailedText("... e disse: Até aqui nos ajudou o Senhor. I Sm 7:12")
		#msg.setButtonText(QMessageBox::Yes,'Sim')
		msg.addButton("OK", QMessageBox.YesRole)
		msg.addButton("All Right", QMessageBox.NoRole)
		msg.addButton("Referências", QMessageBox.ResetRole)
		
		resp = msg.exec_()
		#print(resp)
		if resp == 0:
			print('OK clicado')
		elif resp == 2:
			QMessageBox.about(self, "Referências Bibliográficas", "<b>REFERÊNCIAS</b><hr> <br><br>CHAMBERLAIN PRAVIA, Zacarias M. Projeto e Cálculo de Estruturas de Aço – Edifício Industrial Detalhado. Elsevier. Rio de Janeiro, 2013.<br><br>SOUZA, Alex Sander Clemente de. Dimensionamento de Elementos Estruturais em Aço Segundo a BR 8800:2008. 2009. 109 f. Tese (Doutorado) - Curso de Engenharia Civil, Universidade Federal de São Carlos, São Carlos, 2009. ")
		else:
			QMessageBox.about(self, "Ebenézer", "\''<i>... até aqui o SENHOR nos ajudou'\'.</i>    I Samuel 7.12\n")
			print('All Right clicado')



#---TELA CORTANTE -------------------------------------------------------------------------------
	def calcular_cortante(self):
		cortante = self.lineEdit_66.text()
		dist_enrrijecedor = self.lineEdit_67.text()
		if cortante == ''  or cortante.isdigit() == False :
			QMessageBox.about(self, "Falta de Dados", "Por favor insira todos os dados necessários para o cálculo!\n")
		else:
			cortante = float(self.lineEdit_66.text())
			h = float(self.lineEdit_8.text())/10
			tw = float(self.lineEdit_4.text())/10
			mod_elast = float(self.lineEdit_26.text())
			fy = float(self.lineEdit_24.text())
			d = float(self.lineEdit_2.text())/10

			aw = d*tw
			aw = round(aw,ndigits=4)

			cond_enrijecedor = self.checkBox.isChecked()
			if cond_enrijecedor == True:

				dist_enrrijecedor = float(self.lineEdit_67.text())
				if dist_enrrijecedor !=0 :
					if (dist_enrrijecedor/h > 3.0) or (dist_enrrijecedor/h > ((260/(h/tw))**2)):
						kv = 5.0
					else:
						kv = 5 + (5/(dist_enrrijecedor * (h**2)))
				else:
					kv = 5.0
			else:
				kv = 5.0

			'''if dist_enrrijecedor != '' or dist_enrrijecedor != '0':

				
			else:
				kv = 5.0'''

			kv = round(kv,ndigits=7)

			lambda_cort = h/tw
			lambda_cort = round(lambda_cort,ndigits=3)

			lambda_cort_p = (1.10 * math.sqrt(kv * mod_elast/fy))
			lambda_cort_p = round(lambda_cort_p,ndigits=3)
			
			lambda_cort_r = (1.37 * math.sqrt(kv * mod_elast/fy))
			lambda_cort_r = round(lambda_cort_r,ndigits=3)
			
			vpl = 0.6*aw*fy/10
			vpl = round(vpl,ndigits=3)
			
			#-------------verificação dos casos--------------
			if lambda_cort <= lambda_cort_p:
				print('caso - 1')
				self.label_92.setText('λ <= λp')
				vrd = vpl/1.10
			elif lambda_cort > lambda_cort_p and lambda_cort <= lambda_cort_r:
				print('caso - 2')
				self.label_92.setText('λp < λ <= λr')
				vrd = (lambda_cort_p/lambda_cort)*(vpl/1.10)
			else:
				print('caso - 3')
				self.label_92.setText('λp < λ')
				vrd = 1.24*((lambda_cort_p/lambda_cort)**2)*(vpl/1.10)

			vrd = round(vrd,ndigits=2)

			print('d',d)
			print('tw',tw)
			print('aw',aw)

			self.lineEdit_68.setText(str(lambda_cort))
			self.lineEdit_69.setText(str(lambda_cort_p))
			self.lineEdit_70.setText(str(lambda_cort_r))
			self.lineEdit_73.setText(str(kv))
			self.lineEdit_74.setText(str(aw))
			self.lineEdit_75.setText(str(vpl))
			self.lineEdit_71.setText(str(vrd))

			if float(cortante) <= vrd:
				self.label_188.setText('OK --- Vsd <= Vrd')
				self.label_188.setStyleSheet('color: green')
			else:
				self.label_188.setText('ERRO --- Vsd > Vrd')
				self.label_188.setStyleSheet('color: red')
		

	def alterna_kv(self):
		if self.checkBox.isChecked() == True:
			self.label_82.show()
			self.lineEdit_67.show()
		else:
			self.label_82.hide()
			self.lineEdit_67.hide()

	def limpar_cortante(self):
		self.lineEdit_68.setText('')
		self.lineEdit_69.setText('')
		self.lineEdit_70.setText('')
		self.lineEdit_71.setText('')
		self.lineEdit_73.setText('')
		self.lineEdit_74.setText('')
		self.lineEdit_75.setText('')
		self.lineEdit_73.setText('')

	def salvar_cortante(self):
			QMessageBox.about(self, "Atenção","Sou um Botão sem função implementada. Peça pro Anderson terminar essa função logo, já to entediado :)\n")
#----------------------------------------------------------------------------------



#---TELA FLA -------------------------------------------------------------------------------
	def calcular_fla(self):
		momento = self.lineEdit_38.text()
		comprimt = self.lineEdit_48.text()
		if momento == '' or comprimt == '' or momento.isdigit() == False or comprimt.isdigit() == False :
			None
		else:
			h = float(self.lineEdit_6.text())
			tw = float(self.lineEdit_4.text())
			mod_elast = float(self.lineEdit_26.text())
			fy = float(self.lineEdit_24.text())
			comprimt = float(self.lineEdit_48.text())
			ry = float(self.lineEdit_17.text())
			j_inercia = float(self.lineEdit_19.text())
			wx = float(self.lineEdit_14.text())
			inercia_y = float(self.lineEdit_12.text())
			zx = float(self.lineEdit_16.text())
			cw = float(self.lineEdit_15.text())
			cb = 1.0 #adotei esse valor de cb=1,0 por conta dos exemplos aplicado, se necessário alteral crie uma novo espaço no programa para a entrada desse dado
			wc = wx

			beta = ((fy-0.3*fy)*wx)/(mod_elast*j_inercia)
			beta = round(beta,ndigits=4)

			lambda_fla = h/tw

			lambda_fla_p = (3.76 * math.sqrt(mod_elast/fy))

		
			lambda_fla_p = round(lambda_fla_p,ndigits=3)
			
			lambda_fla_r = (5.70 * math.sqrt(mod_elast/fy))
			lambda_fla_r = round(lambda_fla_r,ndigits=3)
			
			
			mpl = zx * fy
			mr = fy*wx
			mcr = 'procure pelo mcr no anexo H a3!'


			#-------------verificação dos casos--------------
			if lambda_fla <= lambda_fla_p:
				print('caso - 1')
				self.label_78.setText('λ <= λp')
				mrd = mpl/1.10
			else: #lambda_fla > lambda_fla_p and lambda_fla <= lambda_fla_r:
				print('caso - 2')
				self.label_78.setText('λp < λ <= λr')
				mrd = ((1/1.10)*(mpl - (mpl- mr) * ((lambda_fla - lambda_fla_p)/(lambda_fla_r - lambda_fla_p))))
			'''else:
				print('caso - 3')
				self.label_78.setText('λp < λ')
				mrd = mcr/1.10
			'''
			mrd = round(mrd,ndigits=2)
			mpl = round(mpl,ndigits=2)
			mpl = mpl/10
			mr = round(mr,ndigits=2)
			mr = mr/10
			#mcr = round(mcr,ndigits=2)
			#mcr = mcr/100
			mrd = round(mrd,ndigits=2)
			mrd = mrd/10

			print(lambda_fla,' e ',lambda_fla_p,' e ',lambda_fla_r, ' e ', mpl, ' e ',mr, ' e ', mcr, ' e ',beta, ' e ',mrd)
			
			self.lineEdit_58.setText(str(lambda_fla))
			self.lineEdit_59.setText(str(lambda_fla_p))
			self.lineEdit_60.setText(str(lambda_fla_r))
			self.lineEdit_61.setText(str(beta))
			self.lineEdit_62.setText(str(mpl))
			self.lineEdit_63.setText(str(mr))
			self.lineEdit_64.setText(str(mrd))
			#self.lineEdit_65.setText(str(mcr))

			if float(momento) <= mrd:
				self.label_187.setText('OK --- Msd <= Mrd')
				self.label_187.setStyleSheet('color: green')
			else:
				self.label_187.setText('ERRO --- Msd > Mrd')
				self.label_187.setStyleSheet('color: red')

	def limpa_momento(self):
		self.lineEdit_46.setText('')
		self.lineEdit_47.setText('')
		self.lineEdit_39.setText('')
		self.lineEdit_40.setText('')
		self.lineEdit_41.setText('')
		self.lineEdit_42.setText('')
		self.lineEdit_43.setText('')
		self.lineEdit_44.setText('')
		self.lineEdit_45.setText('')
		self.lineEdit_49.setText('')

		self.lineEdit_50.setText('')
		self.lineEdit_51.setText('')
		self.lineEdit_52.setText('')
		self.lineEdit_53.setText('')
		self.lineEdit_54.setText('')
		self.lineEdit_55.setText('')
		self.lineEdit_56.setText('')
		self.lineEdit_57.setText('')

		self.lineEdit_58.setText('')
		self.lineEdit_59.setText('')
		self.lineEdit_60.setText('')
		self.lineEdit_61.setText('')
		self.lineEdit_62.setText('')
		self.lineEdit_63.setText('')
		self.lineEdit_64.setText('')
		self.lineEdit_65.setText('')

	def salvar_momento(self):
			QMessageBox.about(self, "Atenção","Sou um Botão sem função implementada. Peça pro Anderson terminar essa função logo, já to entediado :)\n")

#----------------------------------------------------------------------------------

		

#---TELA FLM -------------------------------------------------------------------------------
	def calcular_flm(self):
		momento = self.lineEdit_38.text()
		comprimt = self.lineEdit_48.text()
		if momento == '' or comprimt == '' or momento.isdigit() == False or comprimt.isdigit() == False :
			None
		else:
			h = float(self.lineEdit_8.text())
			tw = float(self.lineEdit_4.text())
			mod_elast = float(self.lineEdit_26.text())
			fy = float(self.lineEdit_24.text())
			comprimt = float(self.lineEdit_48.text())
			ry = float(self.lineEdit_17.text())
			j_inercia = float(self.lineEdit_19.text())
			wx = float(self.lineEdit_14.text())
			inercia_y = float(self.lineEdit_12.text())
			zx = float(self.lineEdit_16.text())
			cw = float(self.lineEdit_15.text())
			cb = 1.0 #adotei esse valor de cb=1,0 por conta dos exemplos aplicado, se necessário alteral crie uma novo espaço no programa para a entrada desse dado
			wc = wx

			beta = ((fy-0.3*fy)*wx)/(mod_elast*j_inercia)
			beta = round(beta,ndigits=4)

			lambda_flm = float(self.lineEdit_22.text())

			lambda_flm_p = (0.38 * math.sqrt(mod_elast/fy))

		
			lambda_flm_p = round(lambda_flm_p,ndigits=3)
			
			lambda_flm_r = 0.83 * (math.sqrt(mod_elast/(fy - 0.3*fy)))
			lambda_flm_r = round(lambda_flm_r,ndigits=3)
			
			
			mpl = zx * fy
			mr = (fy - 0.3*fy)*wx
			mcr = 0.69*(mod_elast*wc)/(lambda_flm **2)


			#-------------verificação dos casos--------------
			if lambda_flm <= lambda_flm_p:
				print('caso - 1')
				self.label_70.setText('λ <= λp')
				mrd = mpl/1.10
			elif lambda_flm > lambda_flm_p and lambda_flm <= lambda_flm_r:
				print('caso - 2')
				self.label_70.setText('λp < λ <= λr')
				mrd = ((1/1.10)*(mpl - (mpl- mr) * ((lambda_flm - lambda_flm_p)/(lambda_flm_r - lambda_flm_p))))
			else:
				print('caso - 3')
				self.label_70.setText('λp < λ')
				mrd = mcr/1.10


			mpl = round(mpl,ndigits=2)
			mpl = mpl/10
			mr = round(mr,ndigits=2)
			mr = mr/10
			mcr = round(mcr,ndigits=2)
			mcr = mcr/100
			mrd = round(mrd,ndigits=2)
			mrd = mrd/10
			
			self.lineEdit_50.setText(str(lambda_flm))
			self.lineEdit_51.setText(str(lambda_flm_p))
			self.lineEdit_52.setText(str(lambda_flm_r))
			self.lineEdit_55.setText(str(beta))
			self.lineEdit_56.setText(str(mpl))
			self.lineEdit_57.setText(str(mr))
			self.lineEdit_54.setText(str(mcr))
			self.lineEdit_53.setText(str(mrd))

			if float(momento) <= mrd:
				self.label_186.setText('OK --- Msd <= Mrd')
				self.label_186.setStyleSheet('color: green')
			else:
				self.label_186.setText('ERRO --- Msd > Mrd')
				self.label_186.setStyleSheet('color: red')

#----------------------------------------------------------------------------------




#---TELA FLT -------------------------------------------------------------------------------
	def calcular_flt(self):
		momento_flt = self.lineEdit_38.text()
		comprimt = self.lineEdit_48.text()
		if momento_flt == '' or comprimt == '' or momento_flt.isdigit() == False or comprimt.isdigit() == False :
			QMessageBox.about(self, "Falta de Dados", "Por favor insira todos os dados necessários para o cálculo!\n")
		else:
			h = float(self.lineEdit_8.text())
			tw = float(self.lineEdit_4.text())
			mod_elast = float(self.lineEdit_26.text())
			fy = float(self.lineEdit_24.text())
			comprimt = float(self.lineEdit_48.text())
			ry = float(self.lineEdit_17.text())
			j_inercia = float(self.lineEdit_19.text())
			wx = float(self.lineEdit_14.text())
			inercia_y = float(self.lineEdit_12.text())
			zx = float(self.lineEdit_16.text())
			#cw = float(self.lineEdit_15.text())
			cb = float(self.lineEdit_72.text()) #adotei esse valor de cb=1,0 por conta dos exemplos aplicado, se necessário alteral crie uma novo espaço no programa para a entrada desse dado

			d = float(self.lineEdit_2.text())/10
			tf = float(self.lineEdit_7.text())/10
			cw = (inercia_y * ((d - tf)**2)) / 4
			print('cw---------------------',cw)

			fator1 = round((h/tw),ndigits=2)
			fator2 = round(5.7 * math.sqrt(mod_elast/fy),ndigits=2)

			beta = ((fy-0.3*fy)*wx)/(mod_elast*j_inercia)
			beta = round(beta,ndigits=4)
			lamdba_flt = comprimt/ry
			lamdba_flt = round(lamdba_flt,ndigits=3)
			lambda_flt_p = (1.76 * math.sqrt(mod_elast/fy))
			lambda_flt_p = round(lambda_flt_p,ndigits=3)
			lamdba_flt_r = (1.38 * math.sqrt(inercia_y*j_inercia))/(ry*j_inercia*beta) * math.sqrt(1+math.sqrt(1 + ((27*cw*(beta**2))/inercia_y)))
			lamdba_flt_r = round(lamdba_flt_r,ndigits=3)

			mpl = zx * fy
			mr = (fy - 0.3*fy)*wx
			mcr = ((cb*(math.pi **2)*mod_elast*inercia_y)/(comprimt **2)) * (math.sqrt((cw/inercia_y) * (1 + (0.039*j_inercia*(comprimt**2)/cw))))

			

			#-------------verificação dos casos--------------
			if lamdba_flt <= lambda_flt_p:
				print('caso - 1')
				self.label_57.setText('λ <= λp')
				mrd = mpl/1.10

			elif lamdba_flt > lambda_flt_p and lamdba_flt <= lamdba_flt_r:
				print('caso - 2')
				self.label_57.setText('λp < λ <= λr')
				mrd = ((1/1.10)*(mpl - (mpl-mr) * ((lamdba_flt - lambda_flt_p)/(lamdba_flt_r - lambda_flt_p))))
				mrd = cb * mrd

			else:
				print('caso - 3')
				self.label_57.setText('λp < λ')
				mrd = mcr/1.10


			mpl = round(mpl,ndigits=2)
			mpl = mpl/10
			mr = round(mr,ndigits=2)
			mr = mr/10
			mcr = round(mcr,ndigits=2)
			mcr = mcr/10
			mrd = round(mrd,ndigits=2)
			mrd = mrd/10
			
			print(beta,'-',lamdba_flt,'-',lambda_flt_p,'-',lamdba_flt_r)
			self.lineEdit_46.setText(str(fator1))
			self.lineEdit_47.setText(str(fator2))
			self.lineEdit_39.setText(str(lamdba_flt))
			self.lineEdit_40.setText(str(lambda_flt_p))
			self.lineEdit_41.setText(str(lamdba_flt_r))
			self.lineEdit_42.setText(str(beta))
			self.lineEdit_43.setText(str(mpl))
			self.lineEdit_44.setText(str(mr))
			self.lineEdit_49.setText(str(mcr))
			self.lineEdit_45.setText(str(mrd))

			if float(momento_flt) <= mrd:
				self.label_53.setText('OK --- Msd <= Mrd')
				self.label_53.setStyleSheet('color: green')
			else:
				self.label_53.setText('ERRO --- Msd > Mrd')
				self.label_53.setStyleSheet('color: red')

#----------------------------------------------------------------------------------'''


#---TELA NORMAL-------------------------------------------------------------------------------
	def calcular_normal(self):
		#-conferencias---
		normal = self.lineEdit_27.text()
		comprimt = self.lineEdit_36.text()

		if normal == '' or comprimt == '' or normal.isdigit() == False or comprimt.isdigit() == False :
			QMessageBox.about(self, "Falta de Dados", "Por favor insira todos os dados necessários para o cálculo!\n")
		else:
			normal = float(normal)
			comprimt = float(comprimt)
			mod_elast = float(self.lineEdit_26.text())
			fy = float(self.lineEdit_24.text())
			fu = float(self.lineEdit_25.text())
			inercia_x = float(self.lineEdit_18.text())
			inercia_y = float(self.lineEdit_12.text())
			area = float(self.lineEdit_5.text())
			k = float(self.lineEdit_28.text())
			coef_transm =  float(self.lineEdit_37.text())
		

			conf_mesa = round(1.49 * ( math.sqrt(mod_elast/fy) ),ndigits=2)
			conf_alma = round(0.56 * ( math.sqrt(mod_elast/fy) ),ndigits=2)

			#-normal de compressão a flambagem---
			nex = round(((math.pi **2) * mod_elast * inercia_x)/((k*comprimt)**2),ndigits=2)/10
			ney = round(((math.pi **2) * mod_elast * inercia_y)/((k*comprimt)**2),ndigits=2)/10

			lambda0 = math.sqrt((coef_transm*area*fy/10)/(ney))
			lambda0 = round(lambda0,ndigits=2)

			if lambda0 <= 1.5:
				xis = round(0.658 ** (lambda0 **2),ndigits=5)
			else:
				xis = round((0.877/(lambda0 **2)),ndigits=5)

			normal_resit = ((xis*coef_transm*area*fy)/1.1)/10
			normal_resit = round(normal_resit,ndigits=3)

			#print(normal_resit)

			#-retorna os valores para o programa---
			self.lineEdit_30.setText(str(conf_mesa))
			self.lineEdit_29.setText(str(conf_alma))
			self.lineEdit_31.setText(str(nex))
			self.lineEdit_32.setText(str(ney))
			self.lineEdit_33.setText(str(lambda0))
			self.lineEdit_34.setText(str(xis))
			self.lineEdit_35.setText(str(normal_resit))

			if float(normal) <= normal_resit:
				self.label_185.setText('OK --- Vsd <= Vrd')
				self.label_185.setStyleSheet('color: green')
			else:
				self.label_185.setText('ERRO --- Vsd > Vrd')
				self.label_185.setStyleSheet('color: red')


	def altera_K(self):
		tipo_apoio = self.comboBox_3.currentText()
		
		if str(tipo_apoio) == 'Engaste A':
			k = 0.65
		elif str(tipo_apoio) == 'Engaste B':
			k = 0.80
		elif str(tipo_apoio) == 'Engaste C':
			k = 1.20
		elif str(tipo_apoio) == 'Engaste D':
			k = 1.00
		elif str(tipo_apoio) == 'Engaste E':
			k = 2.10
		else:
			k = 2.00
		self.lineEdit_28.setText(str(k))
		#print(tipo_apoio,'-',k)


	def abaEsquemaApoios(self):
		self.dialogo = EsquemaApoios()
		self.dialogo.show()

	def limpa_normal(self):
		self.lineEdit_30.setText('')
		self.lineEdit_29.setText('')
		self.lineEdit_31.setText('')
		self.lineEdit_32.setText('')
		self.lineEdit_33.setText('')
		self.lineEdit_34.setText('')
		self.lineEdit_35.setText('')

	def salvar_normal(self):
			QMessageBox.about(self, "Atenção","Sou um Botão sem função implementada. Peça pro Anderson terminar essa função logo, já to entediado :)\n")

#----------------------------------------------------------------------------------------------------


	def carregar_init_dados(self):
		print('programa inicializado ...') #--------------------------------------------retirar isso futuramente------

		lista_init_perfil = []
		for i in range(1,mcolunas):
			aux_propriedades = planilha.cell(5,i).value
			lista_init_perfil.append(aux_propriedades)

		#print(lista_init_perfil)

		m_linear = float(lista_init_perfil[1])
		self.lineEdit.setText(str(m_linear))

		d = float(lista_init_perfil[2])
		self.lineEdit_2.setText(str(d))

		b = float(lista_init_perfil[3])
		self.lineEdit_3.setText(str(b))

		tw = float(lista_init_perfil[4])
		self.lineEdit_4.setText(str(tw))

		tf = float(lista_init_perfil[5])
		self.lineEdit_7.setText(str(tf))

		d = float(lista_init_perfil[6])
		self.lineEdit_8.setText(str(d))

		d_linha = float(lista_init_perfil[7])
		self.lineEdit_6.setText(str(d_linha))

		area = float(lista_init_perfil[8])
		self.lineEdit_5.setText(str(area))

		ri = float(lista_init_perfil[18])
		self.lineEdit_9.setText(str(ri))

		Inercia = float(lista_init_perfil[19])
		self.lineEdit_19.setText(str(Inercia))

		ix = float(lista_init_perfil[10])
		self.lineEdit_18.setText(str(ix))
		
		wx = float(lista_init_perfil[11])
		self.lineEdit_14.setText(str(wx))
		
		rx = float(lista_init_perfil[12])
		self.lineEdit_11.setText(str(rx))
		
		zx = float(lista_init_perfil[13])
		self.lineEdit_16.setText(str(zx))

		iy = float(lista_init_perfil[14])
		self.lineEdit_12.setText(str(iy))
		
		wy = float(lista_init_perfil[15])
		self.lineEdit_13.setText(str(wy))
		
		ry = float(lista_init_perfil[16])
		self.lineEdit_17.setText(str(ry))
		
		zy = float(lista_init_perfil[17])
		self.lineEdit_10.setText(str(zy))
		
		cw = float(lista_init_perfil[23])
		self.lineEdit_15.setText(str(cw))
		
		u_perimt = float(lista_init_perfil[24])
		self.lineEdit_20.setText(str(u_perimt))

		lamb_f = float(lista_init_perfil[20])
		self.lineEdit_22.setText(str(lamb_f))

		lamb_p = float(lista_init_perfil[21])
		self.lineEdit_23.setText(str(lamb_p))

		lista_init_propriedades_aco = []
		for i in range(1,mcolunas_acos+1):
			aux_propriedades = planilha_acos.cell(4,i).value
			lista_init_propriedades_aco.append(aux_propriedades)

		#print(lista_init_propriedades_aco)

		fy = float(lista_init_propriedades_aco[1])
		self.lineEdit_24.setText(str(fy))

		fu = float(lista_init_propriedades_aco[2])
		self.lineEdit_25.setText(str(fu))

		alongmt = float(lista_init_propriedades_aco[4])
		self.lineEdit_26.setText(str(alongmt))


		self.lineEdit_28.setText('0.65')
		self.lineEdit_37.setText('1.00')

		self.lineEdit_72.setText('1.00')


		self.label_82.hide()
		self.lineEdit_67.hide()


		self.groupBox_19.hide()
		self.lineEdit_65.hide()
		self.label_81.hide()

	def dados_perfil(self):
		perfil = self.comboBox.currentText()
		#print('item: ', perfil)

		posicao = lista.index(perfil)
		#print(posicao)

		lista_propriedades = []
		for i in range(1,mcolunas):
			aux_propriedades = planilha.cell(posicao+5,i).value
			lista_propriedades.append(aux_propriedades)


		print(lista_propriedades)

		tipo_aco = self.comboBox_2.currentText()
		print(tipo_aco)

		m_linear = float(lista_propriedades[1])
		self.lineEdit.setText(str(m_linear))

		d = float(lista_propriedades[2])
		self.lineEdit_2.setText(str(d))

		b = float(lista_propriedades[3])
		self.lineEdit_3.setText(str(b))

		tw = float(lista_propriedades[4])
		self.lineEdit_4.setText(str(tw))

		tf = float(lista_propriedades[5])
		self.lineEdit_7.setText(str(tf))

		d = float(lista_propriedades[6])
		self.lineEdit_8.setText(str(d))

		d_linha = float(lista_propriedades[7])
		self.lineEdit_6.setText(str(d_linha))

		area = float(lista_propriedades[8])
		self.lineEdit_5.setText(str(area))

		ri = float(lista_propriedades[18])
		self.lineEdit_9.setText(str(ri))

		Inercia = float(lista_propriedades[19])
		self.lineEdit_19.setText(str(Inercia))

		ix = float(lista_propriedades[10])
		self.lineEdit_18.setText(str(ix))
		
		wx = float(lista_propriedades[11])
		self.lineEdit_14.setText(str(wx))
		
		rx = float(lista_propriedades[12])
		self.lineEdit_11.setText(str(rx))
		
		zx = float(lista_propriedades[13])
		self.lineEdit_16.setText(str(zx))

		iy = float(lista_propriedades[14])
		self.lineEdit_12.setText(str(iy))
		
		wy = float(lista_propriedades[15])
		self.lineEdit_13.setText(str(wy))
		
		ry = float(lista_propriedades[16])
		self.lineEdit_17.setText(str(ry))
		
		zy = float(lista_propriedades[17])
		self.lineEdit_10.setText(str(zy))
		
		cw = float(lista_propriedades[23])#---------------------------------------------------------------------------------
		self.lineEdit_15.setText(str(cw))
		
		u_perimt = float(lista_propriedades[24])#---------------------------------------------------------------------------------
		self.lineEdit_20.setText(str(u_perimt))

		lamb_f = float(lista_propriedades[20])
		self.lineEdit_22.setText(str(lamb_f))

		lamb_p = float(lista_propriedades[21])
		self.lineEdit_23.setText(str(lamb_p))


	
	def abaEsquemaCb(self):
		self.dialogo = EsquemaCb()
		self.dialogo.show()


	def dados_aco(self):
		aco = self.comboBox_2.currentText()

		posicao = lista_acos.index(aco)
		#print(posicao)

		lista_propriedades_aco = []
		for i in range(1,mcolunas_acos+1):
			aux_propriedades = planilha_acos.cell(posicao+4,i).value
			lista_propriedades_aco.append(aux_propriedades)

		#print(lista_propriedades_aco)

		fy = float(lista_propriedades_aco[1])
		self.lineEdit_24.setText(str(fy))

		fu = float(lista_propriedades_aco[2])
		self.lineEdit_25.setText(str(fu))

		alongmt = float(lista_propriedades_aco[4])
		self.lineEdit_26.setText(str(alongmt))

class EsquemaApoios(QMainWindow):
	def __init__(self):
		super(EsquemaApoios,self).__init__()
		self.setGeometry(50,50,350,183)
		self.setWindowTitle('Condições de Apoio')
		label_img = QtWidgets.QLabel(self)
		label_img.resize(350,183)
		#label_img.setStyleSheet('background-color: red;')
		label_img.setPixmap(QtGui.QPixmap(resource_path('imagens/condicoes_apoio.png')))
		label_img.setScaledContents(True)

		self.setWindowIcon(QtGui.QIcon(resource_path('imagens/brunel_icone.ico')))
		self.setFixedSize(350,183)

class EsquemaCb(QMainWindow):
	def __init__(self):
		super(EsquemaCb,self).__init__()
		self.setGeometry(50,50,350,183)
		self.setWindowTitle('Fator de Modificação')
		label_img = QtWidgets.QLabel(self)
		label_img.resize(280,126)
		#label_img.setStyleSheet('background-color: red;')
		label_img.setPixmap(QtGui.QPixmap(resource_path('imagens/fator_modificacao.png')))
		label_img.setScaledContents(True)

		self.setWindowIcon(QtGui.QIcon(resource_path('imagens/brunel_icone.ico')))
		self.setFixedSize(280,126)





app = QApplication(sys.argv)
a_window = Aplicacao()
sys.exit(app.exec_())