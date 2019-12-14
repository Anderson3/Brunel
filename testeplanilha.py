

from openpyxl import load_workbook

caminho = 'tabela_gerdau.xlsx'
arquivo_excel = load_workbook(caminho)

planilha = arquivo_excel['Plan2']

#print(planilha['a1'].value)

'''mlinhas = planilha.max_row
mcolunas = planilha.max_column

for i in range(1,mlinhas+1):
	for j in range(1,mcolunas+1):
		print(planilha.cell(i,j).value,end='-')'''

mlinhas = planilha.max_row
mcolunas = planilha.max_column
lista = []

for i in range(4,mlinhas+1):
	aux_bitola = planilha.cell(i,1).value
	lista.append(aux_bitola)

print(lista)

'''for i in range(5,mlinhas+1):
	aux_bitola = planilha.cell(i,1).value
	lista.append(aux_bitola)

print(lista)'''